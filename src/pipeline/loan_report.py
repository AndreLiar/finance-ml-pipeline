"""
Loan Decision Report Generator
===============================
Reads the output of creditworthiness.py, retrains the ensemble model,
then calls an LLM to write a professional bank-style loan decision report.

Demonstrates: Deterministic ML (label + numbers) → Probabilistic narrative (LLM prose)

Pipeline position: run AFTER creditworthiness.py
Input : data/creditworthiness_results.xlsx
Output: data/loan_report.txt
        data/creditworthiness_results.xlsx  (new sheet: "Loan Decision Report")

Gap 10: Prompt loaded from prompts/loan_report_v1.txt (git-trackable, version-diffable)
Gap 11: validate_report_numbers() catches hallucinated financial figures
Gap 12: call_llm() wraps ollama — local-only, no data leaves the machine
Gap 13: build_constraints() injects hard constraints derived from ML output into the prompt
"""

import re
import sys
import datetime
import warnings
warnings.filterwarnings('ignore')

import numpy as np
import pandas as pd
from pathlib import Path

from sklearn.preprocessing     import LabelEncoder, StandardScaler
from sklearn.linear_model      import LogisticRegression
from sklearn.ensemble          import RandomForestClassifier, VotingClassifier
from xgboost                   import XGBClassifier

import openpyxl

from src.config import (
    CREDITWORTHINESS_XLSX   as INPUT_EXCEL,
    LOAN_REPORT_TXT         as OUTPUT_TXT,
    CREDITWORTHINESS_XLSX   as OUTPUT_EXCEL,
    LLM_MODEL, LLM_TEMPERATURE, LLM_MAX_TOKENS,
    ROOT,
)
from src.logger import get_logger

log = get_logger(__name__)

# ── Prompt template ────────────────────────────────────────────────────────────
PROMPT_TEMPLATE_PATH = ROOT / "src" / "prompts" / "loan_report_v1.txt"

# Must stay in sync with creditworthiness.py CREDIT_FEATURES
CREDIT_FEATURES = [
    'dscr', 'savings_rate', 'overdraft_freq', 'expense_volatility',
    'income_stability', 'essential_ratio', 'discretionary_ratio',
    'cash_ratio', 'transfer_ratio', 'avg_tx_amount', 'max_tx_amount',
    'tx_count', 'avg_3m_income', 'avg_3m_spend', 'spend_trend',
    'debt_payments',
]

FEATURE_DESCRIPTIONS = {
    'dscr':                'Debt Service Coverage Ratio (income / debt payments)',
    'savings_rate':        'Savings Rate ((income - spend) / income)',
    'overdraft_freq':      'Overdraft Frequency (% of last 6 months in deficit)',
    'expense_volatility':  'Expense Volatility (3-month rolling std of spend, EUR)',
    'income_stability':    'Income Stability (coefficient of variation, lower = stable)',
    'essential_ratio':     'Essential Spend Ratio (groceries/health/insurance/telecom)',
    'discretionary_ratio': 'Discretionary Spend Ratio (restaurants/shopping/travel)',
    'cash_ratio':          'Cash Withdrawal Ratio (% of spend as cash)',
    'transfer_ratio':      'Transfer Ratio (% of spend sent as transfers)',
    'avg_tx_amount':       'Average Transaction Amount (EUR)',
    'max_tx_amount':       'Largest Single Transaction (EUR)',
    'tx_count':            'Transaction Count (per month)',
    'avg_3m_income':       '3-Month Rolling Average Income (EUR)',
    'avg_3m_spend':        '3-Month Rolling Average Spend (EUR)',
    'spend_trend':         'Spend Trend (monthly delta, EUR — positive = increasing)',
    'debt_payments':       'Total Monthly Debt Payments (EUR)',
}

# ── 1. LOAD MONTHLY PROFILE ───────────────────────────────────────────────────

def load_monthly_profile():
    if not INPUT_EXCEL.exists():
        log.error("%s not found. Run creditworthiness.py first.", INPUT_EXCEL)
        sys.exit(1)

    monthly_df = pd.read_excel(INPUT_EXCEL, sheet_name="Monthly Credit Profile")
    monthly_df = monthly_df.sort_values('year_month').reset_index(drop=True)

    latest_row = monthly_df.iloc[-1]
    log.info("Loaded %d months | Most recent: %s", len(monthly_df), latest_row['year_month'])
    return monthly_df, latest_row


# ── 2. RETRAIN MODELS ─────────────────────────────────────────────────────────

def retrain_models(monthly_df):
    le = LabelEncoder()
    X  = monthly_df[CREDIT_FEATURES].values
    y  = le.fit_transform(monthly_df['credit_label'])

    scaler    = StandardScaler()
    X_scaled  = scaler.fit_transform(X)

    lr = LogisticRegression(max_iter=1000, C=1.0, random_state=42, class_weight='balanced')
    rf = RandomForestClassifier(n_estimators=300, max_depth=10, min_samples_leaf=2,
                                random_state=42, class_weight='balanced', n_jobs=-1)
    xgb = XGBClassifier(n_estimators=300, max_depth=4, learning_rate=0.05,
                        subsample=0.8, colsample_bytree=0.8,
                        eval_metric='mlogloss', random_state=42, verbosity=0)

    ensemble = VotingClassifier(
        estimators=[('lr', lr), ('rf', rf), ('xgb', xgb)],
        voting='soft',
        weights=[1, 2, 2],
    )

    rf.fit(X, y)
    ensemble.fit(X_scaled, y)

    fi_df = pd.DataFrame({
        'feature':    CREDIT_FEATURES,
        'importance': rf.feature_importances_,
    }).sort_values('importance', ascending=False).reset_index(drop=True)

    log.info("Ensemble trained | Classes: %s | Top feature: %s (%.4f)",
             list(le.classes_), fi_df.iloc[0]['feature'], fi_df.iloc[0]['importance'])
    return rf, ensemble, scaler, le, fi_df


# ── 3. BUILD PREDICTION CONTEXT ───────────────────────────────────────────────

def build_prediction_context(latest_row, fi_df, ensemble, scaler, le):
    X_latest     = latest_row[CREDIT_FEATURES].values.reshape(1, -1)
    X_latest_s   = scaler.transform(X_latest)

    pred_label_enc = ensemble.predict(X_latest_s)[0]
    pred_proba     = ensemble.predict_proba(X_latest_s)[0]
    pred_label     = le.inverse_transform([pred_label_enc])[0]

    class_proba = {cls: round(prob * 100, 1)
                   for cls, prob in zip(le.classes_, pred_proba)}
    confidence  = class_proba[pred_label]

    top5 = fi_df.head(5).copy()
    top5['value'] = top5['feature'].map(lambda f: latest_row[f])

    positive_signals = []
    risk_signals     = []

    def _sig(condition, pos_msg, risk_msg):
        if condition:
            positive_signals.append(pos_msg)
        else:
            risk_signals.append(risk_msg)

    dscr           = latest_row['dscr']
    savings_rate   = latest_row['savings_rate']
    overdraft_freq = latest_row['overdraft_freq']
    expense_vol    = latest_row['expense_volatility']
    income_stab    = latest_row['income_stability']
    cash_ratio     = latest_row['cash_ratio']
    transfer_ratio = latest_row['transfer_ratio']
    spend_trend    = latest_row['spend_trend']
    disc_ratio     = latest_row['discretionary_ratio']

    _sig(dscr >= 1.5,
         f"DSCR of {dscr:.2f}x meets the minimum 1.5x threshold" + (" (strong)" if dscr >= 3.0 else ""),
         f"DSCR of {dscr:.2f}x is below the minimum 1.5x threshold — income insufficient vs debt obligations")

    _sig(savings_rate >= 0.10,
         f"Savings rate of {savings_rate:.1%} meets the healthy 10%+ benchmark",
         f"Savings rate of {savings_rate:.1%} is below the 10% benchmark" + (" (deficit spending)" if savings_rate < 0 else ""))

    _sig(overdraft_freq == 0,
         "Zero overdraft frequency over the past 6 months",
         f"Overdraft frequency of {overdraft_freq:.1%} over the past 6 months indicates recurring deficits")

    _sig(expense_vol < 300,
         f"Expense volatility of EUR{expense_vol:.0f} indicates stable, predictable spending",
         f"Expense volatility of EUR{expense_vol:.0f} indicates irregular or unpredictable spending patterns")

    _sig(income_stab < 0.3,
         f"Income stability coefficient of {income_stab:.3f} indicates consistent income",
         f"Income stability coefficient of {income_stab:.3f} indicates variable income — a lending risk")

    _sig(spend_trend <= 0,
         f"Spending trend is stable or declining (EUR{spend_trend:+.0f}/month)",
         f"Spending is on an upward trend (EUR{spend_trend:+.0f}/month) — trajectory may worsen affordability")

    _sig(cash_ratio < 0.15,
         f"Low cash withdrawal ratio ({cash_ratio:.1%}) — transactions are traceable",
         f"Cash withdrawal ratio of {cash_ratio:.1%} reduces financial transparency")

    _sig(transfer_ratio < 0.20,
         None,
         f"Transfer ratio of {transfer_ratio:.1%} is elevated — significant funds leaving the account")

    positive_signals = [s for s in positive_signals if s]

    spend_trend_direction = "increasing" if spend_trend > 50 else \
                            "decreasing" if spend_trend < -50 else "stable"

    top_features_lines = []
    for _, row in top5.iterrows():
        desc = FEATURE_DESCRIPTIONS.get(row['feature'], row['feature'])
        val  = row['value']
        imp  = row['importance']
        if row['feature'] in ('savings_rate', 'overdraft_freq', 'essential_ratio',
                               'discretionary_ratio', 'cash_ratio', 'transfer_ratio',
                               'income_stability'):
            val_str = f"{val:.1%}" if row['feature'] != 'income_stability' else f"{val:.3f}"
        elif row['feature'] in ('expense_volatility', 'avg_tx_amount', 'max_tx_amount',
                                 'avg_3m_income', 'avg_3m_spend', 'debt_payments', 'spend_trend'):
            val_str = f"EUR{val:.0f}"
        elif row['feature'] == 'dscr':
            val_str = f"{val:.2f}x"
        else:
            val_str = f"{val:.1f}"
        top_features_lines.append(f"  {row['feature']:<25} importance={imp:.4f}  value={val_str}")

    ctx = {
        'report_date':          datetime.date.today().strftime('%d %B %Y'),
        'analysis_month':       str(latest_row['year_month']),
        'ensemble_label':       pred_label,
        'confidence':           confidence,
        'credit_score':         int(latest_row['credit_score']),
        'stored_label':         str(latest_row['credit_label']),
        'low_risk_pct':         class_proba.get('LOW_RISK', 0.0),
        'medium_risk_pct':      class_proba.get('MEDIUM_RISK', 0.0),
        'high_risk_pct':        class_proba.get('HIGH_RISK', 0.0),
        'dscr':                 round(float(dscr), 2),
        'savings_rate':         round(float(savings_rate), 4),
        'overdraft_freq':       round(float(overdraft_freq), 4),
        'expense_volatility':   round(float(expense_vol), 1),
        'income_stability':     round(float(income_stab), 4),
        'avg_3m_income':        round(float(latest_row['avg_3m_income']), 0),
        'avg_3m_spend':         round(float(latest_row['avg_3m_spend']), 0),
        'spend_trend':          round(float(spend_trend), 1),
        'discretionary_ratio':  round(float(disc_ratio), 4),
        'essential_ratio':      round(float(latest_row['essential_ratio']), 4),
        'cash_ratio':           round(float(cash_ratio), 4),
        'transfer_ratio':       round(float(transfer_ratio), 4),
        'debt_payments':        round(float(latest_row['debt_payments']), 2),
        'avg_tx_amount':        round(float(latest_row['avg_tx_amount']), 2),
        'max_tx_amount':        round(float(latest_row['max_tx_amount']), 2),
        'tx_count':             int(latest_row['tx_count']),
        'top_features_formatted':     '\n'.join(top_features_lines),
        'positive_signals_formatted': '\n'.join(f"  + {s}" for s in positive_signals) or "  (none identified)",
        'risk_signals_formatted':     '\n'.join(f"  - {s}" for s in risk_signals) or "  (none identified)",
        'spend_trend_direction':      spend_trend_direction,
        'n_positive':                 len(positive_signals),
        'n_risk':                     len(risk_signals),
    }

    log.info("Ensemble decision: %s (confidence: %.1f%%)", pred_label, confidence)
    log.info("Credit score: %d/100 | Positive: %d | Risk: %d",
             ctx['credit_score'], ctx['n_positive'], ctx['n_risk'])
    return ctx


# ── Gap 13: BUILD HARD CONSTRAINTS ────────────────────────────────────────────

def build_constraints(ctx: dict) -> str:
    """
    Derive hard constraints from ML output and return a formatted block
    to inject into the prompt. Prevents the LLM from contradicting the ML decision.
    """
    label = ctx['ensemble_label']
    score = ctx['credit_score']
    lines = ["=== HARD CONSTRAINTS (derived from ML output — you MUST follow these) ==="]

    if label == "HIGH_RISK":
        lines.append("- The ML ensemble has classified this applicant as HIGH_RISK.")
        lines.append("- Section 5 FINAL RECOMMENDATION MUST be 'DECLINE' or 'CONDITIONAL APPROVAL'.")
        lines.append("- You MUST NOT write 'APPROVE' as the final recommendation.")
    elif label == "MEDIUM_RISK":
        lines.append("- The ML ensemble has classified this applicant as MEDIUM_RISK.")
        lines.append("- Section 5 FINAL RECOMMENDATION MUST be 'CONDITIONAL APPROVAL' or 'DECLINE'.")
        lines.append("- You MUST NOT write 'APPROVE' (unconditional) as the final recommendation.")
    else:  # LOW_RISK
        lines.append("- The ML ensemble has classified this applicant as LOW_RISK.")
        lines.append("- Section 5 FINAL RECOMMENDATION may be 'APPROVE', 'CONDITIONAL APPROVAL', or 'DECLINE'.")

    if score < 40:
        lines.append(f"- Credit score is {score}/100 (below 40). The narrative must reflect a weak credit profile.")
    elif score >= 70:
        lines.append(f"- Credit score is {score}/100 (above 70). The narrative may acknowledge a relatively strong profile.")

    lines.append("=== END HARD CONSTRAINTS ===\n")
    return '\n'.join(lines)


# ── Gap 10: BUILD PROMPT FROM TEMPLATE FILE ───────────────────────────────────

def build_prompt(ctx: dict) -> str:
    """Load prompt template from file and substitute context values."""
    if not PROMPT_TEMPLATE_PATH.exists():
        log.error("Prompt template not found: %s", PROMPT_TEMPLATE_PATH)
        sys.exit(1)

    template = PROMPT_TEMPLATE_PATH.read_text(encoding='utf-8')
    log.info("Loaded prompt template: %s", PROMPT_TEMPLATE_PATH.name)

    constraints_block = build_constraints(ctx)
    ctx_with_constraints = {**ctx, 'constraints_block': constraints_block}

    try:
        return template.format_map(ctx_with_constraints)
    except KeyError as e:
        log.error("Prompt template references unknown key: %s", e)
        sys.exit(1)


# ── Gap 12: LLM CALL — local Ollama only ─────────────────────────────────────
# This project handles personal financial data. All LLM inference runs locally
# via Ollama (Mistral). No data is sent to external APIs.

def call_llm(prompt_text: str) -> str:
    """Call Mistral via local Ollama. Data never leaves the machine."""
    try:
        import ollama
    except ImportError:
        log.error("'ollama' package not installed. Run: pip install ollama")
        sys.exit(1)

    log.info("LLM: ollama/%s | temperature=%s", LLM_MODEL, LLM_TEMPERATURE)
    try:
        response = ollama.chat(
            model=LLM_MODEL,
            messages=[{"role": "user", "content": prompt_text}],
            options={"temperature": LLM_TEMPERATURE, "num_predict": LLM_MAX_TOKENS},
        )
        report_text = response["message"]["content"]
        log.info("Report generated (%d words)", len(report_text.split()))
        return report_text
    except Exception as e:
        err = str(e)
        if "connection" in err.lower() or "refused" in err.lower():
            log.error("Cannot reach Ollama. Ensure it is running: ollama serve")
            log.error("Ensure Mistral is pulled: ollama pull mistral")
        else:
            log.error("Ollama call failed: %s", err)
        sys.exit(1)


# ── Gap 11: NUMBER VALIDATION ─────────────────────────────────────────────────

# Numeric metric keys and the approximate values we expect to see in the report
_NUMERIC_METRICS = {
    'dscr':               ('dscr',            1.0),
    'savings_rate':       ('savings_rate',     100.0),   # shown as % in prose
    'avg_3m_income':      ('avg_3m_income',    1.0),
    'avg_3m_spend':       ('avg_3m_spend',     1.0),
    'expense_volatility': ('expense_volatility', 1.0),
    'credit_score':       ('credit_score',     1.0),
}

_METRIC_KEYWORDS = {
    'dscr':               ['dscr', 'debt service coverage', 'coverage ratio'],
    'savings_rate':       ['savings rate', 'saving rate'],
    'avg_3m_income':      ['average income', '3-month.*income', 'avg.*income'],
    'avg_3m_spend':       ['average spend', '3-month.*spend', 'avg.*spend'],
    'expense_volatility': ['expense volatility', 'spending volatility'],
    'credit_score':       ['credit score'],
}


def validate_report_numbers(report_text: str, ctx: dict) -> list[str]:
    """
    Extract numbers AFTER known metric keywords and check against ctx values.
    Returns a list of warning strings (empty if all checks pass).

    Only validates metrics with positive expected values — negative rates and
    zero incomes cannot be meaningfully compared as unsigned prose numbers.
    Tolerance is 15% to accommodate LLM rounding.
    """
    warnings_found = []
    text_lower = report_text.lower()

    for metric_key, keywords in _METRIC_KEYWORDS.items():
        ctx_key, scale = _NUMERIC_METRICS[metric_key]
        expected = ctx.get(ctx_key)
        if expected is None:
            continue

        expected_scaled = float(expected) * scale

        # Skip metrics that are zero, negative, or very small —
        # prose representations can't be compared as unsigned numbers
        if expected_scaled <= 0:
            continue

        for kw in keywords:
            pattern = re.compile(kw, re.IGNORECASE)
            for m in pattern.finditer(text_lower):
                # Only look AFTER the keyword (avoid grabbing prior sentence numbers)
                window = report_text[m.end(): m.end() + 120]
                nums = re.findall(r'(\d+(?:\.\d+)?)', window)
                for num_str in nums:
                    found = float(num_str)
                    if found == 0:
                        continue
                    ratio = abs(found - expected_scaled) / expected_scaled
                    if ratio > 0.15:
                        msg = (
                            f"POSSIBLE HALLUCINATION: '{metric_key}' near \"{kw}\" — "
                            f"found {found}, expected ~{expected_scaled:.2f} "
                            f"(deviation {ratio:.0%})"
                        )
                        warnings_found.append(msg)
                        log.warning(msg)
                    break  # only check first number per keyword match
                break  # first keyword match is sufficient

    return warnings_found


# ── POST-PROCESS: VERIFY SECTIONS ─────────────────────────────────────────────

def verify_sections(report_text: str):
    expected = [
        "## 1. EXECUTIVE SUMMARY",
        "## 2. RISK DECISION RATIONALE",
        "## 3. POSITIVE FINANCIAL SIGNALS",
        "## 4. AREAS REQUIRING IMPROVEMENT",
        "## 5. FINAL RECOMMENDATION",
    ]
    missing = [s for s in expected if s not in report_text]
    if missing:
        for s in missing:
            log.warning("Section not found in report: %s", s)
    else:
        log.info("All 5 sections present.")


# ── SAVE REPORT ───────────────────────────────────────────────────────────────

def save_report(report_text: str, ctx: dict, validation_warnings: list[str]):
    # ── Plain text ────────────────────────────────────────────────────────────
    header = (
        f"LOAN DECISION REPORT\n"
        f"{'=' * 60}\n"
        f"Report Date     : {ctx['report_date']}\n"
        f"Analysis Month  : {ctx['analysis_month']}\n"
        f"ML Decision     : {ctx['ensemble_label']} (confidence: {ctx['confidence']:.1f}%)\n"
        f"Credit Score    : {ctx['credit_score']} / 100\n"
        f"Prompt Template : {PROMPT_TEMPLATE_PATH.name}\n"
        f"LLM Backend     : ollama / {LLM_MODEL}\n"
        f"{'=' * 60}\n\n"
        f"NOTE: The decision label and all metrics above are produced by a\n"
        f"deterministic ML ensemble (Logistic Regression + Random Forest +\n"
        f"XGBoost). The narrative below is generated by an LLM and represents\n"
        f"a probabilistic explanation of those fixed outputs.\n"
        f"{'=' * 60}\n\n"
    )

    validation_block = ""
    if validation_warnings:
        validation_block = (
            "\n\n" + "=" * 60 + "\n"
            "VALIDATION WARNINGS — Potential hallucinated numbers detected:\n"
            + "\n".join(f"  ! {w}" for w in validation_warnings)
            + "\n" + "=" * 60
        )

    OUTPUT_TXT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_TXT.write_text(header + report_text + validation_block, encoding='utf-8')
    log.info("Report saved: %s", OUTPUT_TXT)

    # ── Excel sheet ───────────────────────────────────────────────────────────
    wb = openpyxl.load_workbook(OUTPUT_EXCEL)
    if "Loan Decision Report" in wb.sheetnames:
        del wb["Loan Decision Report"]
    ws = wb.create_sheet("Loan Decision Report")

    meta_rows = [
        ("Report Date",      ctx['report_date']),
        ("Analysis Month",   ctx['analysis_month']),
        ("ML Decision",      f"{ctx['ensemble_label']} ({ctx['confidence']:.1f}% confidence)"),
        ("Credit Score",     f"{ctx['credit_score']} / 100"),
        ("Prompt Template",  PROMPT_TEMPLATE_PATH.name),
        ("LLM Backend",      f"ollama / {LLM_MODEL}"),
        ("LOW_RISK prob",    f"{ctx['low_risk_pct']:.1f}%"),
        ("MEDIUM_RISK prob", f"{ctx['medium_risk_pct']:.1f}%"),
        ("HIGH_RISK prob",   f"{ctx['high_risk_pct']:.1f}%"),
        ("", ""),
        ("Section", "Content"),
    ]
    for label, value in meta_rows:
        ws.append([label, value])

    section_pattern = re.compile(r'^(## \d+\..+)$', re.MULTILINE)
    parts = section_pattern.split(report_text)

    def _safe(text):
        _subs = {
            '\u2192': '->', '\u2190': '<-', '\u2191': '^', '\u2193': 'v',
            '\u2022': '-',  '\u2013': '-', '\u2014': '--', '\u2018': "'",
            '\u2019': "'",  '\u201c': '"', '\u201d': '"', '\u2026': '...',
            '\u20ac': 'EUR','\u00e9': 'e', '\u00e8': 'e', '\u00ea': 'e',
        }
        for src, dst in _subs.items():
            text = text.replace(src, dst)
        return text.encode('latin-1', errors='replace').decode('latin-1')

    i = 1
    while i < len(parts):
        section_header = _safe(parts[i].strip())
        section_body   = _safe(parts[i + 1].strip() if i + 1 < len(parts) else "")
        ws.append([section_header, section_body])
        i += 2

    if validation_warnings:
        ws.append(["VALIDATION WARNINGS", "\n".join(validation_warnings)])

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 120
    from openpyxl.styles import Alignment
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    import tempfile, shutil, os
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", dir=OUTPUT_EXCEL.parent)
    os.close(tmp_fd)
    try:
        wb.save(tmp_path)
        shutil.move(tmp_path, OUTPUT_EXCEL)
    except Exception:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        raise
    log.info("Excel updated: %s → sheet 'Loan Decision Report'", OUTPUT_EXCEL)


# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    log.info("=" * 60)
    log.info("LOAN DECISION REPORT GENERATOR")
    log.info("=" * 60)

    log.info("[1/5] Loading monthly credit profile...")
    monthly_df, latest_row = load_monthly_profile()

    log.info("[2/5] Retraining models...")
    rf, ensemble, scaler, le, fi_df = retrain_models(monthly_df)

    log.info("[3/5] Building prediction context for %s...", latest_row['year_month'])
    ctx = build_prediction_context(latest_row, fi_df, ensemble, scaler, le)

    log.info("[4/5] Calling LLM (ollama/%s)...", LLM_MODEL)
    prompt      = build_prompt(ctx)
    report_text = call_llm(prompt)
    verify_sections(report_text)

    log.info("[5/5] Validating numbers and saving outputs...")
    validation_warnings = validate_report_numbers(report_text, ctx)
    if validation_warnings:
        log.warning("%d number validation warning(s) found.", len(validation_warnings))
    else:
        log.info("Number validation passed — no hallucinated figures detected.")

    save_report(report_text, ctx, validation_warnings)

    log.info("=" * 60)
    log.info("DONE")
    log.info("Report  : %s", OUTPUT_TXT)
    log.info("Decision: %s | Score: %d/100", ctx['ensemble_label'], ctx['credit_score'])
    log.info("Breakdown: LOW=%.1f%%  MEDIUM=%.1f%%  HIGH=%.1f%%",
             ctx['low_risk_pct'], ctx['medium_risk_pct'], ctx['high_risk_pct'])


if __name__ == "__main__":
    main()
